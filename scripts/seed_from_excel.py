from __future__ import annotations

import argparse
import sqlite3
import sys
from collections import defaultdict
from statistics import median
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.betting_core import connect, init_db, match_key, normalize_name, rate_tier


DEFAULT_EXCEL = "/Users/ciobez/Downloads/world_cup_2026_betting_model_auto_ready(3) (ultimo).xlsx"


def cell(row, idx):
    return row[idx - 1]


def reset(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DELETE FROM bets;
        DELETE FROM odds;
        DELETE FROM matches;
        DELETE FROM teams;
        """
    )
    conn.commit()


def seed_teams(conn: sqlite3.Connection, wb) -> None:
    ws = wb["Groups"]
    for row in ws.iter_rows(min_row=4, max_row=51, values_only=True):
        group_name, team, fifa_rank, fifa_points, seed_score, strength, _notes = row[:7]
        if not team:
            continue
        conn.execute(
            """
            INSERT OR REPLACE INTO teams(team, normalized_team, group_name, fifa_rank, fifa_points, strength, tier)
            VALUES(?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(team),
                normalize_name(str(team)),
                group_name,
                fifa_rank,
                fifa_points,
                float(strength or 75),
                rate_tier(float(fifa_rank) if fifa_rank else None),
            ),
        )


def seed_matches(conn: sqlite3.Connection, wb) -> None:
    ws = wb["Match_Schedule"]
    for row in ws.iter_rows(min_row=4, max_row=75, values_only=True):
        match_id, group_name, match_date, utc_time, home, away, venue, stage = row[:8]
        if not match_id or not home or not away:
            continue
        conn.execute(
            """
            INSERT OR REPLACE INTO matches(
                match_id, group_name, match_date, utc_time, home_team, away_team, venue, stage, match_key
            )
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(match_id),
                group_name,
                str(match_date),
                str(utc_time or ""),
                normalize_name(str(home)),
                normalize_name(str(away)),
                venue,
                stage,
                match_key(str(home), str(away)),
            ),
        )


def seed_live_odds(conn: sqlite3.Connection, wb) -> None:
    if "Live_Odds" not in wb.sheetnames:
        return
    ws = wb["Live_Odds"]
    grouped = defaultdict(lambda: {"home": [], "draw": [], "away": []})
    for home, away, outcome, price in ws.iter_rows(min_row=2, values_only=True):
        if not home or not away or not outcome or not price:
            continue
        key = match_key(str(home), str(away))
        home_norm = normalize_name(str(home))
        away_norm = normalize_name(str(away))
        outcome_norm = normalize_name(str(outcome))
        try:
            decimal_price = float(price)
        except (TypeError, ValueError):
            continue
        if decimal_price <= 1:
            continue
        if outcome_norm == home_norm:
            grouped[key]["home"].append(decimal_price)
        elif outcome_norm == away_norm:
            grouped[key]["away"].append(decimal_price)
        elif outcome_norm == "Draw":
            grouped[key]["draw"].append(decimal_price)

    for key, values in grouped.items():
        if not all(values[name] for name in ("home", "draw", "away")):
            continue
        conn.execute(
            """
            INSERT INTO odds(match_key, bookmaker, source, home_odds, draw_odds, away_odds, notes)
            VALUES(?, ?, ?, ?, ?, ?, ?)
            """,
            (
                key,
                "Median market",
                "Excel Live_Odds",
                median(values["home"]),
                median(values["draw"]),
                median(values["away"]),
                "Median from workbook rows",
            ),
        )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=DEFAULT_EXCEL)
    parser.add_argument("--keep-bets", action="store_true")
    args = parser.parse_args()

    excel = Path(args.excel)
    if not excel.exists():
        raise SystemExit(f"Excel file not found: {excel}")
    wb = load_workbook(excel, data_only=True, read_only=True)
    conn = connect()
    init_db(conn)
    if args.keep_bets:
        conn.executescript("DELETE FROM odds; DELETE FROM matches; DELETE FROM teams;")
    else:
        reset(conn)
    seed_teams(conn, wb)
    seed_matches(conn, wb)
    seed_live_odds(conn, wb)
    conn.commit()
    print("Seed complete")
    print("teams", conn.execute("SELECT COUNT(*) FROM teams").fetchone()[0])
    print("matches", conn.execute("SELECT COUNT(*) FROM matches").fetchone()[0])
    print("odds", conn.execute("SELECT COUNT(*) FROM odds").fetchone()[0])


if __name__ == "__main__":
    main()
